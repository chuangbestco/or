from __future__ import annotations
import asyncio
import csv, io, json, os, re, time, uuid
from datetime import datetime
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import httpx
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel

ROOT = Path(__file__).resolve().parents[1]
UPLOADS, OUTPUTS = ROOT / 'data' / 'uploads', ROOT / 'outputs'
CONFIG_DIR = Path.home() / '.or-account-info-backfill'
CONFIG_FILE = CONFIG_DIR / 'settings.json'
for folder in (UPLOADS, OUTPUTS, CONFIG_DIR): folder.mkdir(parents=True, exist_ok=True)
try: CONFIG_DIR.chmod(0o700)
except OSError: pass

app = FastAPI(title='OR 账号信息回填')
app.mount('/static', StaticFiles(directory=ROOT / 'app' / 'static'), name='static')
REQUIRED_VALUES = ('account_id', 'ak', 'mk', 'bank_card_tail')
REQUIRED_HEADERS = (*REQUIRED_VALUES, 'register_time', 'register_ip', 'charge_ip', 'remark')
_ADSPOWER_LOCK = asyncio.Lock(); _ADSPOWER_LAST_REQUEST = 0.0
ADSPOWER_MIN_INTERVAL, ADSPOWER_RATE_RETRIES = 1.15, 3
JOBS: dict[str, dict] = {}

class AppError(Exception): pass
class AdsPowerSettings(BaseModel):
    base_url: str
    api_key: str

def norm(v: Any) -> str: return str(v or '').strip()
def norm_header(v: Any) -> str: return norm(v).lower()
def utc_beijing(value: str) -> str: return datetime.fromisoformat(value.replace('Z', '+00:00')).astimezone(ZoneInfo('Asia/Shanghai')).strftime('%y/%m/%d %H:%M:%S')
def tail4(value: Any) -> str:
    digits = ''.join(re.findall(r'\d', norm(value)))
    return digits[-4:].zfill(4) if digits else ''
def read_settings() -> dict | None:
    try:
        data=json.loads(CONFIG_FILE.read_text(encoding='utf-8'))
        return data if norm(data.get('base_url')) and norm(data.get('api_key')) else None
    except (OSError, json.JSONDecodeError): return None
def save_settings(settings: AdsPowerSettings):
    CONFIG_FILE.write_text(json.dumps(settings.model_dump(), ensure_ascii=False), encoding='utf-8')
    try: CONFIG_FILE.chmod(0o600)
    except OSError: pass

def read_csv(raw: bytes) -> tuple[list[str], list[dict[str,str]]]:
    for enc in ('utf-8-sig', 'utf-8', 'gb18030'):
        try:
            reader=csv.DictReader(io.StringIO(raw.decode(enc)))
            return reader.fieldnames or [], list(reader)
        except UnicodeDecodeError: continue
    raise AppError('CSV 编码无法识别，请使用 UTF-8 或 GB18030。')
def read_xlsx(raw: bytes) -> tuple[list[str], list[dict[str,str]]]:
    wb=load_workbook(io.BytesIO(raw), data_only=False, read_only=True)
    for ws in wb.worksheets:
        values=ws.iter_rows(values_only=True); headers=[norm(x) for x in next(values, [])]
        if 'account_id' in [norm_header(x) for x in headers]:
            rows=[{headers[i]:norm(v) for i,v in enumerate(row) if i<len(headers)} for row in values if any(v is not None for v in row)]
            return headers,rows
    raise AppError('XLSX 中未找到含 account_id 表头的工作表。')
def parse_upload(filename: str, raw: bytes):
    ext=Path(filename).suffix.lower()
    if ext=='.csv': return read_csv(raw)
    if ext=='.xlsx': return read_xlsx(raw)
    raise AppError('仅支持 .csv 或 .xlsx 文件。')
def aliases(headers: list[str]) -> dict[str,str]: return {norm_header(h):h for h in headers}
def validation(headers: list[str], rows: list[dict[str,str]]) -> dict:
    cols=aliases(headers); missing=[x for x in REQUIRED_HEADERS if x not in cols]; empty={}
    if not missing:
        for name in REQUIRED_VALUES:
            bad=[str(i+2) for i,row in enumerate(rows) if not norm(row.get(cols[name]))]
            if bad: empty[name]=bad
    ok=bool(rows) and not missing and not empty
    msg='文件字段与实际内容校验成功，可执行回填。' if ok else '文件校验未通过：'
    if missing: msg+='缺少表头：'+'、'.join(missing)+'。'
    if empty: msg+='必填字段为空：'+'；'.join(f'{k} 第{",".join(v)}行' for k,v in empty.items())+'。'
    if not rows: msg+='文件没有数据行。'
    return {'valid':ok,'row_count':len(rows),'headers':headers,'missing_headers':missing,'empty_required_values':empty,'message':msg}

async def adspower_request(client: httpx.AsyncClient, url: str, params: dict, key: str) -> dict:
    global _ADSPOWER_LAST_REQUEST
    async with _ADSPOWER_LOCK:
        for attempt in range(ADSPOWER_RATE_RETRIES):
            delay=ADSPOWER_MIN_INTERVAL-(time.monotonic()-_ADSPOWER_LAST_REQUEST)
            if delay>0: await asyncio.sleep(delay)
            resp=await client.get(url,params=params,headers={'Authorization':f'Bearer {key}'})
            _ADSPOWER_LAST_REQUEST=time.monotonic()
            if resp.status_code!=200: raise AppError(f'AdsPower 连接失败（HTTP {resp.status_code}）。请检查连接地址和 API Key。')
            try: body=resp.json()
            except ValueError as e: raise AppError('AdsPower 返回了无法识别的数据。') from e
            message=norm(body.get('msg') or body.get('message'))
            if body.get('code') in (0,None): return body
            if 'Too many request per second' in message and attempt<ADSPOWER_RATE_RETRIES-1:
                await asyncio.sleep(ADSPOWER_MIN_INTERVAL*(attempt+1)); continue
            raise AppError('AdsPower 返回错误：'+message)
    raise AppError('AdsPower 请求频率受限，请稍后重试。')
async def adspower_profiles(settings: dict) -> list[dict]:
    base=norm(settings['base_url']).rstrip('/')
    if not base.startswith(('http://','https://')): raise AppError('AdsPower 连接地址必须以 http:// 或 https:// 开头。')
    profiles=[]
    async with httpx.AsyncClient(timeout=30,trust_env=False) as client:
        page=1
        while True:
            body=await adspower_request(client,f'{base}/api/v1/user/list',{'page':page,'page_size':100},settings['api_key'])
            batch=(body.get('data') or {}).get('list') or []
            if not isinstance(batch,list): raise AppError('AdsPower 返回格式异常。')
            profiles.extend(batch)
            if len(batch)<100: return profiles
            page+=1

def ip_index(profiles: list[dict]) -> dict[str,list[str]]:
    result={}
    for p in profiles:
        username=norm(p.get('username')).lower(); host=norm((p.get('user_proxy_config') or {}).get('proxy_host')) or norm(p.get('ip'))
        if username and host: result.setdefault(username,[]).append(host)
    return {k:list(dict.fromkeys(v)) for k,v in result.items()}
async def key_metadata(mk: str, client: httpx.AsyncClient | None = None) -> list[dict]:
    async def fetch(c: httpx.AsyncClient) -> list[dict]:
        r=await c.get('https://openrouter.ai/api/v1/keys',headers={'Authorization':f'Bearer {mk}'})
        if r.status_code!=200: raise AppError(f'OpenRouter MK 查询失败（HTTP {r.status_code}）。')
        try: d=r.json()
        except ValueError as e: raise AppError('OpenRouter MK 返回了无法识别的数据。') from e
        data=d.get('data',d.get('keys',d)) if isinstance(d,dict) else d
        return data if isinstance(data,list) else []
    if client: return await fetch(client)
    async with httpx.AsyncClient(timeout=35,trust_env=False) as owned_client:
        return await fetch(owned_client)

async def balance(key: str, client: httpx.AsyncClient | None = None) -> float:
    async def fetch(c: httpx.AsyncClient) -> float:
        r=await c.get('https://openrouter.ai/api/v1/credits',headers={'Authorization':f'Bearer {key}'})
        if r.status_code!=200: raise AppError(f'OpenRouter 余额查询失败（HTTP {r.status_code}）。')
        try: d=r.json().get('data',r.json())
        except ValueError as e: raise AppError('OpenRouter 余额接口返回了无法识别的数据。') from e
        return float(d.get('total_credits',0)-d.get('total_usage',0))
    if client: return await fetch(client)
    async with httpx.AsyncClient(timeout=35,trust_env=False) as owned_client:
        return await fetch(owned_client)
def match(keys: list[dict], ak: str) -> dict|None:
    found=[]
    for k in keys:
        vals=[norm(k.get(x)) for x in ('key','label','hash','name')]
        if any(v==ak or (len(v)>=8 and (ak.startswith(v) or ak.endswith(v) or v.startswith(ak) or v.endswith(ak))) for v in vals): found.append(k)
    unique={json.dumps(x,sort_keys=True):x for x in found}
    return next(iter(unique.values())) if len(unique)==1 else (keys[0] if not found and len(keys)==1 else None)

def job_view(job: dict) -> dict:
    return {k:v for k,v in job.items() if k not in {'_task'}}
def update_job(job:dict, **values): job.update(values)
JOB_CONCURRENCY = 8

async def process_account(row_number: int, row: dict[str, str], cols: dict[str, str], ips: dict[str, list[str]], client: httpx.AsyncClient) -> dict:
    account=norm(row[cols['account_id']]).lower(); ak=norm(row[cols['ak']]); mk=norm(row[cols['mk']])
    result={'row':row_number,'account_id':account,'errors':[],'time_filled':False,'ip_filled':False,'card_normalized':False,'low_balance':None}

    try:
        found=match(await key_metadata(mk,client),ak)
        if not found or not found.get('created_at'): raise AppError('未匹配到唯一 API Key created_at。')
        row[cols['register_time']]=utc_beijing(norm(found['created_at'])); result['time_filled']=True
    except Exception as e:
        result['errors'].append({'stage':'注册时间','error':str(e)})

    try:
        ipvals=ips.get(account,[])
        if not ipvals: raise AppError('未找到对应 AdsPower IP。')
        if len(ipvals)>1: raise AppError('匹配到多个不同 AdsPower IP。')
        row[cols['register_ip']]=ipvals[0]; result['ip_filled']=True
    except AppError as e:
        result['errors'].append({'stage':'注册 IP','error':str(e)})

    try:
        card_tail=tail4(row[cols['bank_card_tail']])
        if not card_tail: raise AppError('银行卡号后四位为空或不含数字。')
        row[cols['bank_card_tail']]=card_tail; result['card_normalized']=True
    except AppError as e:
        result['errors'].append({'stage':'银行卡后四位','error':str(e)})

    try:
        try: remaining=await balance(mk,client)
        except (AppError, httpx.HTTPError, ValueError): remaining=await balance(ak,client)
        if remaining<1: result['low_balance']=remaining
    except Exception as e:
        result['errors'].append({'stage':'余额检查','error':str(e)})
    return result

async def run_job(job: dict, token: str, settings: dict):
    try:
        raw=UPLOADS/f'{token}.bin'; meta=UPLOADS/f'{token}.json'
        filename=json.loads(meta.read_text())['filename']; headers,rows=parse_upload(filename,raw.read_bytes()); check=validation(headers,rows)
        if not check['valid']: raise AppError(check['message'])
        update_job(job,status='running',phase='正在读取 AdsPower 浏览器配置…',total=len(rows),completed=0)
        ips=ip_index(await adspower_profiles(settings)); cols=aliases(headers)
        report={'total_rows':len(rows),'time_filled':0,'ip_filled':0,'card_normalized':0,'errors':[],'low_balance':[]}
        completed=0
        lock=asyncio.Lock()

        async def worker(index: int, row: dict[str, str], client: httpx.AsyncClient):
            nonlocal completed
            result=await process_account(index+2,row,cols,ips,client)
            async with lock:
                completed+=1
                report['time_filled']+=int(result['time_filled'])
                report['ip_filled']+=int(result['ip_filled'])
                report['card_normalized']+=int(result['card_normalized'])
                if result['errors']:
                    report['errors'].append({'row':result['row'],'account_id':result['account_id'],'errors':result['errors']})
                if result['low_balance'] is not None:
                    report['low_balance'].append({'row':result['row'],'account_id':result['account_id'],'balance':result['low_balance']})
                update_job(job,phase=f'正在并发处理账号：已完成 {completed}/{len(rows)}…',completed=completed)

        limits=httpx.Limits(max_connections=JOB_CONCURRENCY,max_keepalive_connections=JOB_CONCURRENCY)
        async with httpx.AsyncClient(timeout=35,trust_env=False,limits=limits) as client:
            await asyncio.gather(*(worker(index,row,client) for index,row in enumerate(rows)))
        out=OUTPUTS/f'{Path(filename).stem}-已回填注册时间IP卡尾.csv'
        with out.open('w',encoding='utf-8-sig',newline='') as f:
            w=csv.DictWriter(f,fieldnames=headers); w.writeheader(); w.writerows(rows)
        with out.open(encoding='utf-8-sig',newline='') as f: saved=list(csv.DictReader(f))
        if len(saved)!=len(rows): raise AppError('输出文件重新读取验证失败。')
        download_url='/api/download/'+out.name
        if report['errors'] or report['low_balance']:
            details=f"回填异常 {len(report['errors'])} 个账号；余额低于 1：{len(report['low_balance'])} 个。已生成 CSV，可下载核对部分回填结果。"
            update_job(job,status='error',phase='任务完成，但存在账号异常',message=details,report=report,download_url=download_url); return
        update_job(job,status='success',phase='处理完成',message='全部账号回填完成，且余额均不低于 1。',report=report,download_url=download_url)
    except Exception as e:
        update_job(job,status='error',phase='任务失败',message=str(e))

@app.get('/')
async def root(): return FileResponse(ROOT/'app/static/index.html')
@app.get('/api/settings')
async def get_settings():
    s=read_settings(); return {'configured':bool(s),'base_url':s['base_url'] if s else ''}
@app.put('/api/settings')
async def put_settings(settings: AdsPowerSettings):
    try:
        profiles=await adspower_profiles(settings.model_dump()); save_settings(settings)
        return {'ok':True,'message':f'AdsPower 连接正常，已保存本机配置（读取 {len(profiles)} 个浏览器配置）。'}
    except (AppError,httpx.HTTPError) as e: return {'ok':False,'message':str(e)}
@app.post('/api/file/validate')
async def validate_file(file: UploadFile=File(...)):
    try:
        raw=await file.read(); headers,rows=parse_upload(file.filename or '',raw); result=validation(headers,rows); token=uuid.uuid4().hex
        (UPLOADS/f'{token}.bin').write_bytes(raw); (UPLOADS/f'{token}.json').write_text(json.dumps({'filename':file.filename},ensure_ascii=False)); result['token']=token
        return result
    except AppError as e: return {'valid':False,'message':str(e),'missing_headers':[],'empty_required_values':{}}
@app.post('/api/process/{token}')
async def process(token: str):
    if not (UPLOADS/f'{token}.bin').exists(): raise HTTPException(404,'上传文件已失效，请重新上传。')
    settings=read_settings()
    if not settings: raise HTTPException(422,'请先保存并校验 AdsPower 连接信息。')
    job_id=uuid.uuid4().hex; job={'id':job_id,'status':'queued','phase':'任务已创建，等待开始…','total':0,'completed':0,'message':'','report':None}
    JOBS[job_id]=job; job['_task']=asyncio.create_task(run_job(job,token,settings))
    return job_view(job)
@app.get('/api/jobs/{job_id}')
async def get_job(job_id: str):
    if job_id not in JOBS: raise HTTPException(404,'任务不存在或已过期。')
    return job_view(JOBS[job_id])
@app.get('/api/download/{name}')
async def download(name: str):
    p=OUTPUTS/Path(name).name
    if not p.exists(): raise HTTPException(404,'文件不存在')
    return FileResponse(p,media_type='text/csv',filename=p.name)
